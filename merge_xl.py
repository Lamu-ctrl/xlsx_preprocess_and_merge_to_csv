import openpyxl
import os
import glob
import time
import csv


class people:
    def __init__(self, name, video=" ", note=" "):
        self.name = name

        self.video_counter = 0
        self.video = video
        if self.video != None:
            self.video_counter += 1

        self.note_counter = 0
        self.note = note
        if self.note != None:
            self.note_counter += 1

    def add_video(self, video_name):
        if self.video is None:
            self.video = ""
        self.video = self.video + '  \n  '+video_name
        self.video_counter += 1

    def add_note(self, note_content):
        if self.note is None:
            self.note = ""
        self.note = self.note + '  \n  '+note_content
        self.note_counter += 1

    def __str__(self):
        return f"{self.name}: {self.video}, {self.note}"

    def output(self):

        return [self.name, self.video_counter, self.video, self.note_counter, self.note]


# 讀取.xlsx檔案


def read_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

# 寫入.xlsx檔案


def write_xlsx(file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in data:
        sheet.append(row)

    workbook.save(file_path)


if __name__ == "__main__":
    # 使用範例
    input_file = "./flod/何馭存期中互評.xlsx"
    output_file = "output.xlsx"
    sel = []

    os.chdir('./flod')
    all_filenames = [i for i in glob.glob('*.xlsx')]
    # print(all_filenames)
    for filename in all_filenames:
        print(filename)
        # data = read_xlsx("./flod/"+filename)
        data = read_xlsx(filename)
        # print(data)
        # print("讀取的資料：", data)
        flag = 0
        for row in data:

            if row[6] != None or row[8] != None:
                flag += 1
                if flag >= 2:
                    sel.append(row)

    print(sel)
    print(len(sel))

    # 已經整合全資料到sel

    all_people = []

    for row in sel:

        # 檢查是否存在名為"Bob"的MyClass實例
        for person in all_people:
            if person.name == row[1]:
                print("存在名為Bob的MyClass實例")
                video = None
                note = None
                if row[6] != None:
                    video = row[6]
                    person.add_video(video)
                if row[8] != None:
                    note = row[8]

                    person.add_note(note)

                break
        else:
            video = None
            note = None
            if row[6] != None:
                video = row[6]
            if row[8] != None:
                note = row[8]
            print("不存在名為Bob的MyClass實例")
            pp = people(row[1], video=video, note=note)
            print(pp)
            all_people.append(pp)
            # time.sleep(10)

    for people in all_people:
        print(people.output())
    os.chdir('..')
    with open('.\output.csv', mode='w+', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        for people in all_people:
            writer.writerow(people.output())

    # write_xlsx(output_file, data)
    # print("已將處理後的資料寫入", output_file)
